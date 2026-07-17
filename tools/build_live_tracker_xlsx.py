#!/usr/bin/env python3
"""Build the ONE-TIME live tracker workbook.

Creates an XLSX that, once uploaded to Google Drive (converted to a Google
Sheet), pulls tracker/prospects.csv from GitHub raw via IMPORTDATA. The
resulting Google Sheet is permanent: every `git push` of the tracker CSV
refreshes it automatically. Only run this again if the CSV URL changes, the
columns change, or the summary / conditional formatting changes (then
re-import so the new layout takes effect).
"""
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_URL = (
    "https://raw.githubusercontent.com/twbsportsguy/ClaudeCode/"
    "main/tracker/prospects.csv"
)

NAVY = "13294B"          # Finley wordmark navy
BRAND_BLUE = "9FC2E6"    # Finley pine-sprig / pinecone light blue
LIGHT_BLUE = "E8F1F8"
# Drop the club logo here (PNG, transparent bg) and it's embedded on the
# Summary tab automatically; otherwise a styled wordmark stands in.
LOGO_PATH = Path(__file__).resolve().parent.parent / "assets" / "finley-logo.png"
GREEN_FILL = "C8E6C9"
GREEN_TEXT = "1B5E20"
AMBER_FILL = "FFF3CD"
AMBER_TEXT = "7A5C00"
GRAY_FILL = "ECEFF1"
GRAY_TEXT = "455A64"

# Pipeline stages (value stored in the Status column) -> fill color.
# Order matters for the summary funnel (least to most committed).
STAGES = [
    ("Interested: 50%", 0.50, "FFF9C4"),    # light yellow
    ("Red-Hots: 75%", 0.75, "F8BBD0"),       # pink
    ("Agreements: 90%", 0.90, "BBDEFB"),     # light blue
    ("Signed: 100%", 1.00, "C8E6C9"),        # light green
    ("Not Interested: 0%", 0.00, "FFE0B2"),  # light orange
]

CURRENCY_FMT = '"$"#,##0'

# CSV column order -> sheet columns A..W
#  A Date Added   B Rank      C Score      D Company       E Industry
#  F City         G State     H Website    I Revenue       J Employees
#  K Mktg Budget  L Ad Spend  M Why Rank   N Contact Name  O Contact Title
#  P Email        Q Phone     R ZI Co ID   S Draft Created T Status
#  U Next Step    V Notes     W Potential Revenue
WIDTHS = {
    "A": 11, "B": 6, "C": 6, "D": 32, "E": 14, "F": 13, "G": 6, "H": 26,
    "I": 9, "J": 10, "K": 14, "L": 42, "M": 40, "N": 19, "O": 24, "P": 30,
    "Q": 14, "R": 12, "S": 7, "T": 18, "U": 28, "V": 34, "W": 16,
}
LAST_COL = 23  # W
LAST_COL_LETTER = get_column_letter(LAST_COL)
MAX_ROWS = 300


def build_prospects(ws):
    ws.freeze_panes = "E2"
    ws["A1"] = f'=IMPORTDATA("{CSV_URL}")'
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 26
    for c in range(1, LAST_COL + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for r in range(2, MAX_ROWS):
        for col in ("L", "M", "U", "V", "O"):
            ws[f"{col}{r}"].alignment = Alignment(vertical="top", wrap_text=True)
        for col in ("B", "C", "G", "S"):
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"W{r}"].alignment = Alignment(horizontal="right", vertical="top")
        ws[f"W{r}"].number_format = CURRENCY_FMT

    body = f"A2:{LAST_COL_LETTER}{MAX_ROWS}"
    ws.conditional_formatting.add(body, FormulaRule(
        formula=['AND($D2<>"",ISEVEN(ROW()))'], stopIfTrue=False,
        fill=PatternFill("solid", fgColor=LIGHT_BLUE)))
    rank = f"B2:B{MAX_ROWS}"
    ws.conditional_formatting.add(rank, FormulaRule(
        formula=['$B2="A"'], stopIfTrue=True,
        fill=PatternFill("solid", fgColor=GREEN_FILL),
        font=Font(bold=True, color=GREEN_TEXT)))
    ws.conditional_formatting.add(rank, FormulaRule(
        formula=['$B2="B"'], stopIfTrue=True,
        fill=PatternFill("solid", fgColor=AMBER_FILL),
        font=Font(bold=True, color=AMBER_TEXT)))
    ws.conditional_formatting.add(rank, FormulaRule(
        formula=['$B2="C"'], stopIfTrue=True,
        fill=PatternFill("solid", fgColor=GRAY_FILL),
        font=Font(bold=True, color=GRAY_TEXT)))
    ws.conditional_formatting.add(f"S2:S{MAX_ROWS}", FormulaRule(
        formula=['$S2="Y"'], stopIfTrue=True,
        font=Font(bold=True, color=GREEN_TEXT)))

    # Pipeline stage color-coding on the Status column (T). Exact-match each
    # stage string so "Not Interested: 0%" never trips the "Interested" rule.
    status = f"T2:T{MAX_ROWS}"
    for label, _prob, fill in STAGES:
        ws.conditional_formatting.add(status, FormulaRule(
            formula=[f'$T2="{label}"'], stopIfTrue=True,
            fill=PatternFill("solid", fgColor=fill),
            font=Font(bold=True, color="222222")))


def _brand_header(ws):
    """Finley branding band across the top of the Summary sheet: the club
    logo if assets/finley-logo.png exists, else a styled wordmark, plus a
    light-blue accent rule."""
    ws.row_dimensions[1].height = 60
    logo_added = False
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            ratio = (img.width / img.height) if img.height else 2.0
            img.height = 68
            img.width = int(68 * ratio)
            ws.add_image(img, "B1")
            logo_added = True
        except Exception:
            logo_added = False
    if not logo_added:
        ws.merge_cells("B1:H1")
        ws["B1"] = "FINLEY  ·  GOLF CLUB"
        ws["B1"].font = Font(bold=True, size=22, color=NAVY)
        ws["B1"].alignment = Alignment(vertical="center")
    # Light-blue accent rule under the header band.
    ws.row_dimensions[2].height = 5
    for c in range(2, 9):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=BRAND_BLUE)


def build_summary(ws):
    ws.sheet_view.showGridLines = False
    _brand_header(ws)
    ws.merge_cells("B3:H3")
    ws["B3"] = "Corporate Partnerships Pipeline"
    ws["B3"].font = Font(bold=True, size=18, color=NAVY)
    ws.merge_cells("B4:H4")
    ws["B4"] = 'Tyler Baity · live view — updates automatically from the prospect tracker'
    ws["B4"].font = Font(size=11, color="666666")

    kpis = [
        ("Companies", '=COUNTUNIQUE(Prospects!D2:D)', NAVY),
        ("Contacts", "=COUNTA(Prospects!N2:N)", NAVY),
        ("Drafts Ready", '=COUNTIF(Prospects!S2:S,"Y")', GREEN_TEXT),
        ("A Companies", '=COUNTUNIQUE(IFERROR(FILTER(Prospects!D2:D,Prospects!B2:B="A")))', GREEN_TEXT),
        ("B Companies", '=COUNTUNIQUE(IFERROR(FILTER(Prospects!D2:D,Prospects!B2:B="B")))', AMBER_TEXT),
        ("C Companies", '=COUNTUNIQUE(IFERROR(FILTER(Prospects!D2:D,Prospects!B2:B="C")))', GRAY_TEXT),
    ]
    for i, (label, formula, color) in enumerate(kpis):
        col = 2 + i
        v, l = ws.cell(row=5, column=col), ws.cell(row=6, column=col)
        v.value, l.value = formula, label
        v.font = Font(bold=True, size=22, color=color)
        l.font = Font(size=10, color="666666")
        v.alignment = l.alignment = Alignment(horizontal="center")

    # --- Pipeline funnel (counts + potential + probability-weighted revenue) ---
    ws["B9"] = "Pipeline by stage"
    ws["B9"].font = Font(bold=True, size=12, color=NAVY)
    hdr = ["Stage", "Contacts", "Potential $", "Weighted $"]
    for j, text in enumerate(hdr):
        cell = ws.cell(row=10, column=2 + j)
        cell.value = text
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="left" if j == 0 else "center")

    row = 11
    for label, prob, fill in STAGES:
        s = ws.cell(row=row, column=2)  # B: Stage
        s.value = label
        s.fill = PatternFill("solid", fgColor=fill)
        s.font = Font(bold=True, color="222222")
        ws.cell(row=row, column=3).value = f'=COUNTIF(Prospects!T2:T,"{label}")'
        ws.cell(row=row, column=4).value = f'=SUMIF(Prospects!T2:T,"{label}",Prospects!W2:W)'
        ws.cell(row=row, column=5).value = f'={prob}*SUMIF(Prospects!T2:T,"{label}",Prospects!W2:W)'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5).number_format = CURRENCY_FMT
        row += 1

    # Totals row (open pipeline = everything except Not Interested).
    first, last = 11, row - 1
    t = ws.cell(row=row, column=2)
    t.value = "Open pipeline"
    t.font = Font(bold=True, color=NAVY)
    ws.cell(row=row, column=3).value = f'=SUM(C{first}:C{last})-C{last}'  # exclude Not Interested count
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3).font = Font(bold=True, color=NAVY)
    ws.cell(row=row, column=4).value = f'=SUM(D{first}:D{last-1})'  # potential, excl. Not Interested
    ws.cell(row=row, column=5).value = f'=SUM(E{first}:E{last})'    # weighted (Not Interested weights to 0 anyway)
    for c in (4, 5):
        ws.cell(row=row, column=c).number_format = CURRENCY_FMT
        ws.cell(row=row, column=c).font = Font(bold=True, color=NAVY)

    # --- Latest activity ---
    act = row + 2
    ws.cell(row=act, column=2).value = "Latest activity (most recent first)"
    ws.cell(row=act, column=2).font = Font(bold=True, size=12, color=NAVY)
    ws.cell(row=act + 1, column=2).value = (
        '=QUERY(Prospects!A2:W,"select A, B, D, N, T, U where D is not null '
        'order by A desc, B asc limit 15 '
        "label A 'Date', B 'Rank', D 'Company', N 'Contact', T 'Status', U 'Next Step'\",1)"
    )
    for c in range(2, 8):
        cell = ws.cell(row=act + 1, column=c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color="FFFFFF")

    for col, w in {"B": 22, "C": 12, "D": 28, "E": 20, "F": 14, "G": 30, "H": 12}.items():
        ws.column_dimensions[col].width = w


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Sales Tracker (live).xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    build_summary(ws)
    build_prospects(wb.create_sheet("Prospects"))
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
