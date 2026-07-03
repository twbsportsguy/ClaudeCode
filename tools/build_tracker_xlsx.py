#!/usr/bin/env python3
"""Build a presentation-ready XLSX from tracker/prospects.csv.

Usage: python3 tools/build_tracker_xlsx.py [output.xlsx]

Uploading the XLSX to Google Drive (with conversion enabled) yields a
Google Sheet that keeps all of this formatting.
"""
import csv
import sys
from collections import Counter, OrderedDict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
CSV_PATH = REPO / "tracker" / "prospects.csv"

NAVY = "13294B"        # Carolina navy
CAROLINA = "7BAFD4"    # Carolina blue
LIGHT_BLUE = "E8F1F8"  # banding
GREEN = "2E7D32"
AMBER = "F9A825"
GRAY = "9E9E9E"

RANK_FILL = {
    "A": PatternFill("solid", fgColor=GREEN),
    "B": PatternFill("solid", fgColor=AMBER),
    "C": PatternFill("solid", fgColor=GRAY),
}
RANK_FONT = {
    "A": Font(bold=True, color="FFFFFF"),
    "B": Font(bold=True, color="13294B"),
    "C": Font(bold=True, color="FFFFFF"),
}

# Presentation column order (keys = CSV headers)
COLUMNS = OrderedDict([
    ("Rank", 7), ("Score", 7), ("Company", 34), ("City", 14), ("State", 7),
    ("Contact Name", 20), ("Contact Title", 26), ("Contact Email", 32),
    ("Contact Phone", 15), ("Industry", 16), ("Revenue", 10),
    ("Employees", 11), ("Marketing Budget", 15), ("Ad Spend Signals", 46),
    ("Why This Rank", 42), ("Status", 10), ("Draft Created", 8),
    ("Next Step", 30), ("Date Added", 12), ("Website", 30), ("Notes", 40),
])
WRAP_COLS = {"Ad Spend Signals", "Why This Rank", "Next Step", "Notes", "Contact Title"}

thin = Side(style="thin", color="D0D7E2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def load_rows():
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_summary(ws, rows):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:H2")
    t = ws["B2"]
    t.value = "Finley Golf Club — Corporate Partnerships Pipeline"
    t.font = Font(bold=True, size=18, color=NAVY)
    ws.merge_cells("B3:H3")
    s = ws["B3"]
    s.value = f"Prospect tracker · updated {date.today():%B %-d, %Y} · Tyler Baity"
    s.font = Font(size=11, color="666666")

    companies = {r["Company"] for r in rows}
    contacts = sum(1 for r in rows if r["Contact Name"].strip())
    drafts = sum(1 for r in rows if r["Draft Created"].strip().upper() == "Y")
    ranks = Counter(r["Rank"] for r in {r["Company"]: r for r in rows}.values())

    kpis = [
        ("Companies", len(companies), NAVY),
        ("Contacts", contacts, NAVY),
        ("Drafts Ready", drafts, GREEN),
        ("A Prospects", ranks.get("A", 0), GREEN),
        ("B Prospects", ranks.get("B", 0), AMBER),
        ("C Prospects", ranks.get("C", 0), GRAY),
    ]
    for i, (label, val, color) in enumerate(kpis):
        col = 2 + i
        c1, c2 = ws.cell(row=5, column=col), ws.cell(row=6, column=col)
        c1.value, c2.value = val, label
        c1.font = Font(bold=True, size=22, color=color)
        c2.font = Font(size=10, color="666666")
        c1.alignment = c2.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.cell(row=9, column=2, value="Company").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=9, column=3, value="Rank").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=9, column=4, value="Score").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=9, column=5, value="Contacts").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=9, column=6, value="Drafts").font = Font(bold=True, color="FFFFFF")
    ws.merge_cells(start_row=9, start_column=6, end_row=9, end_column=8)
    for col in range(2, 9):
        ws.cell(row=9, column=col).fill = PatternFill("solid", fgColor=NAVY)

    by_company = OrderedDict()
    for r in rows:
        by_company.setdefault(r["Company"], []).append(r)
    ordered = sorted(by_company.items(), key=lambda kv: (kv[1][0]["Rank"], -int(kv[1][0]["Score"] or 0)))
    for i, (name, crows) in enumerate(ordered):
        row = 10 + i
        rank = crows[0]["Rank"]
        ws.cell(row=row, column=2, value=name)
        rc = ws.cell(row=row, column=3, value=rank)
        rc.fill = RANK_FILL.get(rank, PatternFill())
        rc.font = RANK_FONT.get(rank, Font())
        rc.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=int(crows[0]["Score"] or 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=sum(1 for r in crows if r["Contact Name"].strip())).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=sum(1 for r in crows if r["Draft Created"].strip().upper() == "Y")).alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            for col in range(2, 9):
                ws.cell(row=row, column=col).fill = RANK_FILL[rank] if col == 3 else PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.column_dimensions["B"].width = 40


def build_prospects(ws, rows):
    headers = list(COLUMNS)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = COLUMNS[h]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    rows = sorted(rows, key=lambda r: (r["Rank"], -int(r["Score"] or 0), r["Company"]))
    for ri, r in enumerate(rows, start=2):
        band = ri % 2 == 0
        for ci, h in enumerate(headers, start=1):
            val = r.get(h, "")
            if h == "Score" and val:
                val = int(val)
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = BORDER
            c.alignment = Alignment(
                vertical="top",
                horizontal="center" if h in {"Rank", "Score", "State", "Draft Created", "Employees"} else "left",
                wrap_text=h in WRAP_COLS,
            )
            if band:
                c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            if h == "Rank":
                c.fill = RANK_FILL.get(r["Rank"], c.fill)
                c.font = RANK_FONT.get(r["Rank"], c.font)
            elif h == "Company":
                c.font = Font(bold=True, color=NAVY)
            elif h == "Draft Created":
                c.font = Font(bold=True, color=GREEN if str(val).upper() == "Y" else GRAY)


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else REPO / "tracker" / "Sales Tracker.xlsx"
    rows = load_rows()
    wb = Workbook()
    build_summary(wb.active, rows)
    wb.active.title = "Summary"
    build_prospects(wb.create_sheet("Prospects"), rows)
    wb.save(out)
    print(f"Wrote {out} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
